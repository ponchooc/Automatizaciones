# ®Carlos Alfonso Ortega Molina®
# script para procesamiento de reporte petota con integracion wms

import ftplib
import gzip
import importlib.util
import json
import os
import re
import shutil
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

import pandas as pd
import requests
from openpyxl import load_workbook

try:
    import paramiko
except Exception:
    paramiko = None

try:
    from log_control import Logger, Control
except ModuleNotFoundError:
    log_control_code = '''# Utilidad para log y control\nimport os\nfrom datetime import datetime\nimport json\n\nclass Logger:\n    def __init__(self, carpeta):\n        fecha = datetime.now().strftime('%d%m%Y')\n        self.log_path = os.path.join(carpeta, f'log_{fecha}.txt')\n    def log(self, mensaje):\n        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')\n        linea = f'[{ts}] {mensaje}\\n'\n        with open(self.log_path, 'a', encoding='utf-8') as f:\n            f.write(linea)\n        print(linea.strip())\n\nclass Control:\n    def __init__(self, carpeta):\n        fecha = datetime.now().strftime('%d%m%Y')\n        self.ctrl_path = os.path.join(carpeta, f'control_{fecha}.json')\n        self.data = self._cargar()\n    def _cargar(self):\n        if os.path.exists(self.ctrl_path):\n            try:\n                with open(self.ctrl_path, 'r', encoding='utf-8') as f:\n                    return json.load(f)\n            except Exception:\n                return {}\n        return {}\n    def guardar(self):\n        with open(self.ctrl_path, 'w', encoding='utf-8') as f:\n            json.dump(self.data, f, ensure_ascii=False, indent=2)\n    def registrar(self, int_art, resultado):\n        self.data[int_art] = resultado\n        self.guardar()\n    def consultados(self):\n        return set(self.data.keys())\n    def terminado(self, total):\n        return len(self.data) >= total\n    def ultimos(self, n=5):\n        return list(self.data.items())[-n:]\n'''
    with open(os.path.join(os.getcwd(), 'log_control.py'), 'w', encoding='utf-8') as f:
        f.write(log_control_code)
    from log_control import Logger, Control


CONFIG_FILE = ".acceso_gnx"
CONFIG_WMS = ".acceso_wms"

FTP_HOST = "140.240.11.1"
FTP_DIR_REPORTE = "/reportes_gnx"
FTP_DIR_RELICES = "/gnx_prod/manto/desa/trabajo/sears/carlos_ortega"

FECHA_DMY = datetime.now().strftime('%d%m%Y')
FECHA_YMD = datetime.now().strftime('%Y-%m-%d')
RUN_DIR = os.path.join(os.getcwd(), f"run_{FECHA_DMY}")

ARCHIVO_REMOTO_FTP = "scns_petota_anx.csv.gz"
ARCHIVO_RELICES_REMOTO = "relices.txt"
ARCHIVO_RELICES_PROCESADO_REMOTO = "relices_procesado.txt"

ARCHIVO_LOCAL_GZ = os.path.join(RUN_DIR, "scns_petota_anx.csv.gz")
ARCHIVO_LOCAL_CSV = os.path.join(RUN_DIR, "scns_petota.csv")
ARCHIVO_RELICES = os.path.join(RUN_DIR, "relices.txt")
ARCHIVO_RELICES_PROCESADO = os.path.join(RUN_DIR, "relices_procesado.txt")
ARCHIVO_SALIDA = os.path.join(RUN_DIR, "Petota_con_Existencia_WMS.xlsx")

BASE_URL_WMS = "https://e6.wms.ocs.oraclecloud.com/sears2/wms/lgfapi/v10/report/custom_inventory_summary/"
BASE_URL_ORDER_HDR = "https://e6.wms.ocs.oraclecloud.com/sears2/wms/lgfapi/v10/entity/order_hdr/"

STATUS_MAP = {
    0: "Creada",
    10: "Parcialmente Asignado",
    20: "Asignado",
    25: "En Preparacion",
    27: "Recogido",
    30: "En Empaquetado",
    40: "Empaquetado",
    50: "Cargado",
    90: "Enviado",
    99: "Cancelado",
}

# ============================================================================
# configuracion envio correo nueva funcionalidad
# ============================================================================
# flag para activar desactivar envio automatico de reporte por correo
ENVIAR_CORREO_ACTIVO = True  # cambiar a false para desactivar esta funcionalidad

# flag para ejecutar solo la fase de envio de correo saltando todo lo demas
# util cuando el proceso completo ya corrio y solo quieres enviar el correo
SOLO_ENVIO_CORREO = False  # cambiar a true para solo envio ssh correo

# directorio ftp para archivo comprimido del reporte
FTP_DIR_SOLO_PETOTA = "/gnx_prod/manto/desa/trabajo/sears/carlos_ortega/solo_petota"

# Destinatarios del correo (separados por coma)
CORREO_DESTINATARIOS = "ortegac@sanborns.com.mx,noverona@sanborns.com.mx,l.alvarez@soyldm.com,c.arzate@soyldm.com,l.granados@soyldm.com,y.garcia@soyldm.com,jlopez@sears.com.mx,secretariasc@sears.com.mx,mramirezv@sears.com.mx,rgodinez@sears.com.mx,mdominguezc@sears.com.mx,mesaoracle@sears.com.mx,daperez@sears.com.mx,moralesj@sanborns.com.mx,lisalas@sears.com.mx"

# Asunto del correo
CORREO_ASUNTO = "Reporte Petota con WMS"

# Cuerpo del correo
CORREO_CUERPO = "Buen dia, Se adjunta el reporte petota Saludos"

# ============================================================================
# CONFIGURACION RETRY SSH
# ============================================================================
SSH_RETRY_ACTIVO = True
SSH_RETRY_MAX_INTENTOS = 3
SSH_RETRY_ESPERAS_SEG = [30, 90]

# Para correo se usa modo conservador para evitar posibles duplicados
SSH_RETRY_CORREO_MAX_INTENTOS = 1
# ============================================================================


def hoy_dmy():
    return FECHA_DMY


def hoy_ymd():
    return FECHA_YMD


def nombre_archivo_comprimido():
    """genera nombre para archivo comprimido petota_con_existencia_wms_ddmm.zip"""
    ddmm = FECHA_DMY[:4]  # extrae dd y mm de fecha_dmy ejemplo 27022026 da 2702
    return f"Petota_con_Existencia_WMS_{ddmm}.zip"


def hidden_checkpoint_name(nombre):
    return os.path.join(RUN_DIR, f".ckpt_{nombre}_{hoy_dmy()}.ok")


def write_hidden_checkpoint(nombre, contenido="ok"):
    path = hidden_checkpoint_name(nombre)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(contenido)
    return path


def exists_hidden_checkpoint(nombre):
    return os.path.exists(hidden_checkpoint_name(nombre))


def mirror_console_file_path():
    return os.path.join(RUN_DIR, f".console_{hoy_dmy()}.txt")


def phase_header(bitacora, titulo):
    line = "╔" + "═" * 78 + "╗"
    mid = f"║ {titulo:<76}║"
    end = "╚" + "═" * 78 + "╝"
    log_console(bitacora, line)
    log_console(bitacora, mid)
    log_console(bitacora, end)


def cargar_credenciales(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"No se encuentra el archivo {path}")
    with open(path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    user = config.get("usuario")
    password = config.get("contrasena")
    if not user or not password:
        raise ValueError(f"Credenciales inválidas en {path}")
    return user, password


def log_only(logger, mensaje):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    linea = f"[{ts}] {mensaje}"
    with open(logger.log_path, 'a', encoding='utf-8') as f:
        f.write(linea + "\n")
    with open(mirror_console_file_path(), 'a', encoding='utf-8') as f:
        f.write(linea + "\n")


def log_console(logger, mensaje):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    linea = f"[{ts}] {mensaje}"
    print(linea, flush=True)
    with open(logger.log_path, 'a', encoding='utf-8') as f:
        f.write(linea + "\n")
    with open(mirror_console_file_path(), 'a', encoding='utf-8') as f:
        f.write(linea + "\n")


def buscar_archivo_ftp(ftp, nombre):
    archivos = ftp.nlst()
    return nombre if nombre in archivos else None


def ftp_descargar(logger, remote_dir, remote_name, local_name):
    log_console(logger, f"FTP DESCARGA: iniciando conexión a {FTP_HOST}")
    user, password = cargar_credenciales(CONFIG_FILE)
    ftp = ftplib.FTP(FTP_HOST, encoding='latin-1')
    ftp.set_debuglevel(2)
    log_console(logger, f"FTP DESCARGA: conectado. Login con usuario={user}")
    ftp.login(user, password)
    log_console(logger, f"FTP DESCARGA: cambiando directorio a {remote_dir}")
    ftp.cwd(remote_dir)
    log_console(logger, f"FTP DESCARGA: listando archivos en {remote_dir}")
    archivo = buscar_archivo_ftp(ftp, remote_name)
    if not archivo:
        log_console(logger, f"FTP DESCARGA: archivo no encontrado -> {remote_name}")
        ftp.quit()
        log_console(logger, "FTP DESCARGA: conexión cerrada")
        return False

    total_bytes = {'n': 0}
    last_tick = {'t': time.time()}

    def write_chunk(chunk):
        f.write(chunk)
        total_bytes['n'] += len(chunk)
        now = time.time()
        if now - last_tick['t'] >= 1.0:
            log_console(logger, f"FTP DESCARGA: recibidos {total_bytes['n']} bytes...")
            last_tick['t'] = now

    log_console(logger, f"FTP DESCARGA: iniciando RETR {archivo}")
    with open(local_name, 'wb') as f:
        ftp.retrbinary(f'RETR {archivo}', write_chunk)
    log_console(logger, f"FTP DESCARGA: transferencia completada ({total_bytes['n']} bytes)")
    ftp.quit()
    log_console(logger, "FTP DESCARGA: conexión cerrada")
    log_console(logger, f"Descargado por FTP: {remote_name} -> {local_name}")
    return True


def ftp_subir(logger, local_name, remote_dir, remote_name):
    log_console(logger, f"FTP SUBIDA: iniciando conexión a {FTP_HOST}")
    user, password = cargar_credenciales(CONFIG_FILE)
    ftp = ftplib.FTP(FTP_HOST, encoding='latin-1')
    ftp.set_debuglevel(2)
    log_console(logger, f"FTP SUBIDA: conectado. Login con usuario={user}")
    ftp.login(user, password)
    log_console(logger, f"FTP SUBIDA: cambiando directorio a {remote_dir}")
    ftp.cwd(remote_dir)

    total_bytes = {'n': 0}
    last_tick = {'t': time.time()}

    def read_chunk(chunk):
        total_bytes['n'] += len(chunk)
        now = time.time()
        if now - last_tick['t'] >= 1.0:
            log_console(logger, f"FTP SUBIDA: enviados {total_bytes['n']} bytes...")
            last_tick['t'] = now

    log_console(logger, f"FTP SUBIDA: iniciando STOR {remote_name}")
    with open(local_name, 'rb') as f:
        ftp.storbinary(f'STOR {remote_name}', f, callback=read_chunk)
    log_console(logger, f"FTP SUBIDA: transferencia completada ({total_bytes['n']} bytes)")
    ftp.quit()
    log_console(logger, "FTP SUBIDA: conexión cerrada")
    log_console(logger, f"Subido por FTP: {local_name} -> {remote_dir}/{remote_name}")


def descomprimir_gz(logger):
    if not os.path.exists(ARCHIVO_LOCAL_GZ):
        raise FileNotFoundError(f"No existe {ARCHIVO_LOCAL_GZ}")
    with gzip.open(ARCHIVO_LOCAL_GZ, 'rb') as f_in:
        with open(ARCHIVO_LOCAL_CSV, 'wb') as f_out:
            shutil.copyfileobj(f_in, f_out)
    log_only(logger, f"Archivo descomprimido: {ARCHIVO_LOCAL_CSV}")


def leer_csv(logger):
    if not os.path.exists(ARCHIVO_LOCAL_CSV):
        raise FileNotFoundError(f"No existe {ARCHIVO_LOCAL_CSV}")

    def es_error_fila_corrupta(error):
        msg = str(error)
        return 'Error tokenizing data' in msg or ('Expected ' in msg and ' fields in line ' in msg)

    ultimo_error = None
    for enc in ['utf-8', 'latin1']:
        for sep in ['\t', ',']:
            try:
                df = pd.read_csv(ARCHIVO_LOCAL_CSV, sep=sep, encoding=enc)
                if sep == '\t' and len(df.columns) == 1:
                    continue
                return df
            except Exception as e:
                ultimo_error = e
                if not es_error_fila_corrupta(e):
                    continue

                try:
                    log_console(
                        logger,
                        f"LECTURA CSV: se detectaron filas corruptas con sep={repr(sep)} y encoding={enc}. Se reintentará omitiéndolas. Detalle: {str(e)[:200]}"
                    )
                    df = pd.read_csv(
                        ARCHIVO_LOCAL_CSV,
                        sep=sep,
                        encoding=enc,
                        engine='python',
                        on_bad_lines='skip'
                    )
                    if sep == '\t' and len(df.columns) == 1:
                        continue
                    log_console(logger, "LECTURA CSV: archivo leído omitiendo filas corruptas.")
                    return df
                except Exception as skip_error:
                    ultimo_error = skip_error
    raise RuntimeError(f"No se pudo leer CSV: {ultimo_error}")


def consultar_item_wms(sku, auth):
    item_code = f"SRS{sku}"
    params = {
        "facility_id__code": "VAL",
        "company_id__code": "GPOSAN",
        "item_code": item_code,
        "write_header_line_flg": "true"
    }
    try:
        r = requests.get(BASE_URL_WMS, params=params, auth=auth, timeout=25)
        if r.status_code == 404:
            return sku, 0, 'NO ENCONTRADO'
        if r.status_code == 200:
            lineas = r.text.strip().split('\n')
            if len(lineas) > 1:
                headers = lineas[0].split('|')
                datos = lineas[1].split('|')
                d = dict(zip(headers, datos))
                v1 = int(d.get('active_available', 0) or 0)
                v2 = int(d.get('iblpn_total', 0) or 0)
                v3 = int(d.get('iblpn_lockcode', 0) or 0)
                existencia_wms = v1 + v2 - v3
                return sku, existencia_wms, 'OK'
            return sku, 0, 'NO ENCONTRADO'
        return sku, 0, f'ERROR API {r.status_code}'
    except requests.exceptions.Timeout:
        return sku, 0, 'TIEMPO AGOTADO'
    except Exception as e:
        return sku, 0, f'ERROR: {str(e)[:80]}'


def normalize_sales_check(v):
    digits = re.sub(r'\D', '', str(v or ''))
    if not digits:
        return ''
    if len(digits) > 16:
        digits = digits[-16:]
    return digits.zfill(16)


def strip_leading_zero(s):
    x = str(s or '').lstrip('0')
    return x if x else '0'


def crear_relices_txt(logger, df):
    if 'sales check' not in df.columns:
        raise RuntimeError("No existe columna 'sales check'")

    vals = [normalize_sales_check(v) for v in df['sales check'].tolist()]
    vals = [v for v in vals if v]
    uniq = list(dict.fromkeys(vals))

    with open(ARCHIVO_RELICES, 'w', encoding='utf-8') as f:
        for v in uniq:
            f.write(v + '\n')

    with open(ARCHIVO_RELICES, 'r', encoding='utf-8') as f:
        lines = [x.strip() for x in f.readlines() if x.strip()]

    if len(lines) != len(uniq):
        raise RuntimeError("Error de verificación en relices.txt")

    log_console(logger, f"sales check únicos normalizados a 16 dígitos: {len(uniq)}")
    log_only(logger, f"relices.txt verificado con {len(lines)} registros")
    return uniq


def ssh_send_char_by_char(shell, comando, delay=0.1):
    for ch in comando:
        shell.send(ch)
        time.sleep(delay)
    shell.send('\n')


def ssh_drain_output(shell, logger, etiqueta, timeout_seg=2.0, poll_seg=0.2):
    fin = time.time() + timeout_seg
    recibido = False
    while time.time() < fin:
        if shell.recv_ready():
            data = shell.recv(65535)
            if not data:
                break
            txt = data.decode('utf-8', errors='ignore')
            if txt:
                for linea in txt.splitlines():
                    if linea.strip():
                        log_console(logger, f"SSH {etiqueta}: {linea}")
                recibido = True
        else:
            time.sleep(poll_seg)
    return recibido


def ssh_send_and_log(shell, logger, comando, espera_salida=2.0):
    log_console(logger, f"SSH CMD >>> {comando}")
    ssh_send_char_by_char(shell, comando, 0.1)
    ssh_drain_output(shell, logger, "OUT", timeout_seg=espera_salida)


def ejecutar_con_retry_ssh(logger, nombre_fase, accion, max_intentos, esperas_seg):
    if max_intentos <= 1:
        return accion()

    ultimo_error = None
    for intento in range(1, max_intentos + 1):
        try:
            log_console(logger, f"{nombre_fase}: intento {intento}/{max_intentos}")
            return accion()
        except Exception as e:
            ultimo_error = e
            if intento >= max_intentos:
                break

            idx_espera = min(intento - 1, len(esperas_seg) - 1) if esperas_seg else 0
            espera = esperas_seg[idx_espera] if esperas_seg else 30
            log_console(logger, f"{nombre_fase}: fallo intento {intento}: {str(e)[:160]}")
            log_console(logger, f"{nombre_fase}: reintentando en {espera} segundos...")
            time.sleep(espera)

    raise RuntimeError(f"{nombre_fase}: falló tras {max_intentos} intentos. Último error: {ultimo_error}")


def _ejecutar_menu_gnx_once(logger):
    if paramiko is None:
        raise RuntimeError("Paramiko no está disponible. Instala paramiko para usar SSH interactivo")

    user, password = cargar_credenciales(CONFIG_FILE)
    log_console(logger, f"SSH: iniciando conexión a {FTP_HOST} con usuario={user}")

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(
        hostname=FTP_HOST,
        username=user,
        password=password,
        timeout=30,
        look_for_keys=False,
        allow_agent=False,
    )
    log_console(logger, "SSH: conexión establecida")
    shell = client.invoke_shell()
    log_console(logger, "SSH: shell interactivo abierto")
    time.sleep(2)
    ssh_drain_output(shell, logger, "OUT", timeout_seg=2.0)

    ssh_send_and_log(shell, logger, '1', espera_salida=2.0)
    ssh_send_and_log(shell, logger, '1', espera_salida=2.0)
    ssh_send_and_log(shell, logger, f'cd {FTP_DIR_RELICES}', espera_salida=1.5)
    ssh_send_and_log(shell, logger, './relizador.sh', espera_salida=3.0)

    log_console(logger, "SSH: monitoreando salida de relizador.sh en tiempo real...")
    for _ in range(60):
        ssh_drain_output(shell, logger, "OUT", timeout_seg=0.5, poll_seg=0.1)
        time.sleep(0.5)

    ssh_send_and_log(shell, logger, 'exit', espera_salida=1.5)
    ssh_send_and_log(shell, logger, 'f', espera_salida=1.5)

    log_console(logger, "SSH: cerrando shell")
    shell.close()
    log_console(logger, "SSH: cerrando conexión")
    client.close()
    log_console(logger, "Secuencia SSH interactiva GNX ejecutada")


def ejecutar_menu_gnx(logger):
    if SSH_RETRY_ACTIVO:
        return ejecutar_con_retry_ssh(
            logger,
            "SSH GNX",
            lambda: _ejecutar_menu_gnx_once(logger),
            SSH_RETRY_MAX_INTENTOS,
            SSH_RETRY_ESPERAS_SEG,
        )
    return _ejecutar_menu_gnx_once(logger)


def esperar_y_descargar_relices_procesado(logger, max_espera_seg=1800, cada_seg=30):
    inicio = time.time()
    while True:
        log_console(logger, f"FTP DESCARGA ESPERA: intentando descargar {ARCHIVO_RELICES_PROCESADO}...")
        ok = ftp_descargar(logger, FTP_DIR_RELICES, ARCHIVO_RELICES_PROCESADO_REMOTO, ARCHIVO_RELICES_PROCESADO)
        if ok:
            log_console(logger, f"Descargado {ARCHIVO_RELICES_PROCESADO}")
            return True
        if time.time() - inicio > max_espera_seg:
            log_console(logger, f"FTP DESCARGA ESPERA: tiempo agotado para {ARCHIVO_RELICES_PROCESADO}")
            return False
        log_console(logger, f"{ARCHIVO_RELICES_PROCESADO} aún no disponible. Reintentando en {cada_seg}s...")
        time.sleep(cada_seg)


def parsear_relices_procesado(logger):
    if not os.path.exists(ARCHIVO_RELICES_PROCESADO):
        raise FileNotFoundError(f"No existe {ARCHIVO_RELICES_PROCESADO}")

    mapeo = {}
    with open(ARCHIVO_RELICES_PROCESADO, 'r', encoding='utf-8', errors='ignore') as f:
        for linea in f:
            line = linea.strip()
            if not line:
                continue
            if '|' in line:
                parts = [p.strip() for p in line.split('|')]
            elif '\t' in line:
                parts = [p.strip() for p in line.split('\t')]
            elif ',' in line:
                parts = [p.strip() for p in line.split(',')]
            else:
                parts = [line]
            if len(parts) < 2:
                continue
            sales = normalize_sales_check(parts[0].strip())
            order_release = parts[1].strip()
            if sales and order_release:
                mapeo[sales] = order_release

    log_console(logger, f"relices_procesado parseado. registros válidos: {len(mapeo)}")
    return mapeo


def consultar_order_status_lote(order_nbr_list, auth):
    if not order_nbr_list:
        return {}, []

    pedidos = [str(x).strip() for x in order_nbr_list if str(x).strip()]
    params = {
        'facility_id__code': 'VAL',
        'company_id__code': 'GPOSAN',
        'fields': 'order_nbr,status_id',
        'order_nbr__in': ','.join(pedidos),
    }
    try:
        r = requests.get(BASE_URL_ORDER_HDR, params=params, auth=auth, timeout=25)
        if r.status_code != 200:
            # ERROR API 404 equivale al estatus de SIN ENVIO A WMS.
            if r.status_code == 404:
                return {o: "SIN ENVIO A WMS" for o in pedidos}, []
            return {o: f"ERROR API {r.status_code}" for o in pedidos}, []

        payload = r.json()
        results = payload.get('results', [])

        by_order = {}
        for row in results:
            order_nbr = str(row.get('order_nbr', '')).strip()
            if not order_nbr:
                continue
            status_id = row.get('status_id')
            if status_id is None:
                by_order[order_nbr] = 'No Encontrada'
            else:
                try:
                    by_order[order_nbr] = STATUS_MAP.get(int(status_id), f"Estatus {status_id}")
                except Exception:
                    by_order[order_nbr] = f"Estatus {status_id}"

        final_map = {}
        for order_nbr in pedidos:
            final_map[order_nbr] = by_order.get(order_nbr, 'No Encontrada')

        return final_map, []
    except requests.exceptions.Timeout:
        return {o: 'TIEMPO AGOTADO' for o in pedidos}, list(pedidos)
    except Exception as e:
        return {o: f"ERROR: {str(e)[:80]}" for o in pedidos}, []


def consultar_order_status_unitario(order_nbr, auth):
    order_nbr = str(order_nbr).strip()
    params = {
        'facility_id__code': 'VAL',
        'company_id__code': 'GPOSAN',
        'fields': 'order_nbr,status_id',
        'order_nbr': order_nbr,
    }
    try:
        r = requests.get(BASE_URL_ORDER_HDR, params=params, auth=auth, timeout=25)
        if r.status_code != 200:
            # ERROR API 404 equivale al estatus de SIN ENVIO A WMS.
            if r.status_code == 404:
                return "SIN ENVIO A WMS"
            return f"ERROR API {r.status_code}"
        payload = r.json()
        results = payload.get('results', [])
        if not results:
            return 'No Encontrada'

        status_id = results[0].get('status_id')
        if status_id is None:
            return 'No Encontrada'
        try:
            return STATUS_MAP.get(int(status_id), f"Estatus {status_id}")
        except Exception:
            return f"Estatus {status_id}"
    except requests.exceptions.Timeout:
        return 'TIEMPO AGOTADO'
    except Exception as e:
        return f"ERROR: {str(e)[:80]}"


def guardar_excel(df, logger):
    illegal_re = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
    for col in df.select_dtypes(include=['object', 'string']).columns:
        df[col] = df[col].apply(lambda v: illegal_re.sub('', str(v)) if v is not None else "")

    for col in ['sales check', 'ean_art', 'telefono', 'Existencia WMS']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype('int64')

    if not importlib.util.find_spec('openpyxl'):
        raise RuntimeError("openpyxl no disponible")

    df.to_excel(ARCHIVO_SALIDA, index=False, engine='openpyxl')

    wb = load_workbook(ARCHIVO_SALIDA)
    ws = wb.active

    for column_cells in ws.columns:
        max_len = 0
        for cell in column_cells:
            val = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 65)

    for col_name in ['sales check', 'ean_art', 'telefono', 'Existencia WMS']:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = '0'

    wb.save(ARCHIVO_SALIDA)
    return ARCHIVO_SALIDA





def _ejecutar_enviador_correos_ssh_once(logger, nombre_excel, nombre_comprimido):
    """ejecuta la secuencia de compresion envio y limpieza via ssh
    
    pasos automatizados en servidor remoto
    1 comprime el archivo excel usando zip en modo sin directorios
    2 ejecuta el script enviador_correos.sh con el archivo comprimido
    3 elimina ambos archivos excel y comprimido del servidor remoto
    """
    if paramiko is None:
        raise RuntimeError("Paramiko no está disponible. Instala paramiko para usar SSH interactivo")

    user, password = cargar_credenciales(CONFIG_FILE)
    log_console(logger, f"SSH CORREO: iniciando conexión a {FTP_HOST} con usuario={user}")

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(
        hostname=FTP_HOST,
        username=user,
        password=password,
        timeout=30,
        look_for_keys=False,
        allow_agent=False,
    )
    log_console(logger, "SSH CORREO: conexión establecida")
    shell = client.invoke_shell()
    log_console(logger, "SSH CORREO: shell interactivo abierto")
    time.sleep(2)
    ssh_drain_output(shell, logger, "CORREO", timeout_seg=2.0)

    # ®Carlos Alfonso Ortega Molina®
    # navegacion al directorio de trabajo en servidor remoto
    ssh_send_and_log(shell, logger, '1', espera_salida=2.0)
    ssh_send_and_log(shell, logger, '1', espera_salida=2.0)
    ssh_send_and_log(shell, logger, f'cd {FTP_DIR_SOLO_PETOTA}', espera_salida=1.5)
    
    # paso 1 compresion del archivo excel usando formato zip
    # flag j comprime sin incluir estructura de directorios
    log_console(logger, f"SSH CORREO: comprimiendo {nombre_excel} a {nombre_comprimido}...")
    comando_compresion = f'zip -j {nombre_comprimido} {nombre_excel}'
    ssh_send_and_log(shell, logger, comando_compresion, espera_salida=2.0)
    
    # paso 2 ejecucion del script de envio con archivo comprimido como adjunto
    comando_correo = f'./enviador_correos.sh {CORREO_DESTINATARIOS} "{CORREO_ASUNTO}" "{CORREO_CUERPO}" {nombre_comprimido}'
    log_console(logger, f"SSH CORREO: ejecutando script de envío")
    log_console(logger, f"  Destinatarios: {CORREO_DESTINATARIOS}")
    log_console(logger, f"  Asunto: {CORREO_ASUNTO}")
    log_console(logger, f"  Adjunto comprimido: {nombre_comprimido}")
    ssh_send_and_log(shell, logger, comando_correo, espera_salida=3.0)

    # monitoreo de salida del script en tiempo real
    log_console(logger, "SSH CORREO: monitoreando salida del script enviador_correos.sh...")
    for _ in range(60):
        ssh_drain_output(shell, logger, "CORREO", timeout_seg=0.5, poll_seg=0.1)
        time.sleep(0.5)

    # paso 3 limpieza eliminacion de archivos temporales del servidor
    # flag f fuerza eliminacion sin confirmar evita errores si no existen
    log_console(logger, f"SSH CORREO: eliminando archivos del servidor...")
    comando_limpieza = f'rm -f {nombre_excel} {nombre_comprimido}'
    ssh_send_and_log(shell, logger, comando_limpieza, espera_salida=1.5)
    log_console(logger, f"SSH CORREO: archivos eliminados: {nombre_excel} y {nombre_comprimido}")

    # cierre de sesion ssh
    ssh_send_and_log(shell, logger, 'exit', espera_salida=1.5)
    ssh_send_and_log(shell, logger, 'f', espera_salida=1.5)

    log_console(logger, "SSH CORREO: cerrando shell")
    shell.close()
    log_console(logger, "SSH CORREO: cerrando conexión")
    client.close()
    log_console(logger, "SSH CORREO: Secuencia de compresión, envío y limpieza ejecutada exitosamente")


def ejecutar_enviador_correos_ssh(logger, nombre_excel, nombre_comprimido):
    if SSH_RETRY_ACTIVO and SSH_RETRY_CORREO_MAX_INTENTOS > 1:
        return ejecutar_con_retry_ssh(
            logger,
            "SSH CORREO",
            lambda: _ejecutar_enviador_correos_ssh_once(logger, nombre_excel, nombre_comprimido),
            SSH_RETRY_CORREO_MAX_INTENTOS,
            SSH_RETRY_ESPERAS_SEG,
        )
    log_console(logger, "SSH CORREO: modo conservador de retry activo (1 intento para evitar duplicados de envío)")
    return _ejecutar_enviador_correos_ssh_once(logger, nombre_excel, nombre_comprimido)


def ejecutar_fase_envio_correo(logger, control, hoy, archivo_final):
    # ========================================================================
    # fase opcional envio reporte por correo con compresion zip en servidor
    # ejecuta subida ftp compresion con zip envio de correo y limpieza remota
    # ========================================================================
    if ENVIAR_CORREO_ACTIVO:
        phase_header(logger, "ENVIO REPORTE POR CORREO")

        correo_phase_done = (
            control.data.get('_correo_phase_done') is True
            and control.data.get('_correo_phase_done_fecha') == hoy
        )

        if not correo_phase_done:
            # nombres de archivos local excel y remoto zip
            nombre_excel = os.path.basename(archivo_final)
            nombre_comprimido = nombre_archivo_comprimido()  # petota_con_existencia_wms_ddmm.zip
            
            # paso 1 subir archivo excel via ftp al servidor remoto
            correo_ftp_done = (
                control.data.get('_correo_ftp_done') is True
                and control.data.get('_correo_ftp_done_fecha') == hoy
            )

            if not correo_ftp_done:
                if not exists_hidden_checkpoint('correo_ftp_subido'):
                    log_console(logger, f"CORREO: Subiendo archivo Excel {nombre_excel} vía FTP...")
                    log_console(logger, f"  Origen local: {archivo_final}")
                    log_console(logger, f"  Destino FTP: {FTP_DIR_SOLO_PETOTA}/{nombre_excel}")
                    ftp_subir(logger, archivo_final, FTP_DIR_SOLO_PETOTA, nombre_excel)
                    write_hidden_checkpoint('correo_ftp_subido', 'ok')
                else:
                    log_console(logger, "Checkpoint detectado: FTP correo ya subido.")

                control.data['_correo_ftp_done'] = True
                control.data['_correo_ftp_done_fecha'] = hoy
                control.guardar()
            else:
                log_console(logger, "Fase FTP correo ya completada hoy.")

            # paso 2 ejecutar compresion zip envio y limpieza via ssh
            correo_ssh_done = (
                control.data.get('_correo_ssh_done') is True
                and control.data.get('_correo_ssh_done_fecha') == hoy
            )

            if not correo_ssh_done:
                if not exists_hidden_checkpoint('correo_ssh_ejecutado'):
                    log_console(logger, "CORREO: Ejecutando secuencia de compresión, envío y limpieza vía SSH...")
                    log_console(logger, f"  Archivo Excel: {nombre_excel}")
                    log_console(logger, f"  Archivo comprimido ZIP: {nombre_comprimido}")
                    ejecutar_enviador_correos_ssh(logger, nombre_excel, nombre_comprimido)
                    write_hidden_checkpoint('correo_ssh_ejecutado', 'ok')
                else:
                    log_console(logger, "Checkpoint detectado: SSH correo ya ejecutado.")

                control.data['_correo_ssh_done'] = True
                control.data['_correo_ssh_done_fecha'] = hoy
                control.guardar()
            else:
                log_console(logger, "Fase SSH correo ya completada hoy.")

            # marcado de fase como completa con timestamp
            control.data['_correo_phase_done'] = True
            control.data['_correo_phase_done_fecha'] = hoy
            control.data['_correo_phase_done_ts'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            control.guardar()
            log_console(logger, "Fase ENVIO CORREO marcada como COMPLETADA para hoy.")
        else:
            log_console(logger, "Fase ENVIO CORREO ya completada hoy. Se omite.")
    else:
        log_console(logger, "ENVIO CORREO DESACTIVADO (ENVIAR_CORREO_ACTIVO = False)")


def main():
    os.makedirs(RUN_DIR, exist_ok=True)
    logger = Logger(RUN_DIR)
    control = Control(RUN_DIR)

    # limpieza de claves temporales legacy de versiones anteriores
    temp_keys_legacy = ['_inventario_temp_limit', '_inventario_temp_ts', '_order_temp_limit', '_order_temp_ts']
    cleaned = False
    for key in temp_keys_legacy:
        if key in control.data:
            control.data.pop(key, None)
            cleaned = True
    if cleaned:
        control.guardar()
        log_console(logger, "Se limpiaron marcas temporales heredadas del control.")

    phase_header(logger, "INICIO PROCESO PETOTA GNX")

    hoy = hoy_ymd()

    # modo especial solo envio de correo salta toda la logica principal
    # util para reenviar correo cuando el proceso ya ejecuto completamente
    if SOLO_ENVIO_CORREO:
        phase_header(logger, "MODO SOLO ENVIO CORREO ACTIVADO")
        if not os.path.exists(ARCHIVO_SALIDA):
            raise FileNotFoundError(f"No existe archivo Excel para enviar: {ARCHIVO_SALIDA}")
        log_console(logger, f"Archivo Excel encontrado: {ARCHIVO_SALIDA}")
        log_console(logger, "Ejecutando unicamente fase de envio de correo...")
        ejecutar_fase_envio_correo(logger, control, hoy, ARCHIVO_SALIDA)
        log_only(logger, f"=== PROCESO FINALIZADO SOLO CORREO: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
        return

    # verificacion de fases ya completadas para omitir trabajo duplicado
    excel_done_hoy = (
        control.data.get('_excel_phase_done') is True
        and control.data.get('_excel_phase_done_fecha') == hoy
        and os.path.exists(ARCHIVO_SALIDA)
    )
    correo_done_hoy = (
        control.data.get('_correo_phase_done') is True
        and control.data.get('_correo_phase_done_fecha') == hoy
    )

    # logica de reinicio inteligente excel ya generado solo falta correo
    if excel_done_hoy:
        if ENVIAR_CORREO_ACTIVO and not correo_done_hoy:
            log_console(logger, "Excel ya generado hoy y envío de correo pendiente. Se ejecutará únicamente la fase de correo.")
            ejecutar_fase_envio_correo(logger, control, hoy, ARCHIVO_SALIDA)
            log_only(logger, f"=== PROCESO FINALIZADO CORRECTAMENTE: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
            return
        log_console(logger, "Proceso ya finalizado hoy. Se omiten todas las fases.")
        return

    descargado_hoy = control.data.get('_descargado_fecha') == hoy
    descomprimido_hoy = control.data.get('_descomprimido_fecha') == hoy

    if not descargado_hoy or not os.path.exists(ARCHIVO_LOCAL_GZ):
        phase_header(logger, "DESCARGA FTP SCNS")
        exito = False
        for intento, espera in enumerate([0, 600, 300], 1):
            if espera > 0:
                log_only(logger, f"Esperando {espera//60} minutos antes del intento {intento}...")
                time.sleep(espera)
            if ftp_descargar(logger, FTP_DIR_REPORTE, ARCHIVO_REMOTO_FTP, ARCHIVO_LOCAL_GZ):
                exito = True
                control.data['_descargado_fecha'] = hoy
                control.guardar()
                break
        if not exito:
            raise RuntimeError("No se pudo descargar scns_petota_anx.csv.gz")
    else:
        log_console(logger, "ETAPA 2: Archivo ya descargado hoy. Se omite descarga.")

    if not descomprimido_hoy or not os.path.exists(ARCHIVO_LOCAL_CSV):
        phase_header(logger, "DESCOMPRESION SCNS")
        descomprimir_gz(logger)
        control.data['_descomprimido_fecha'] = hoy
        control.guardar()
    else:
        log_console(logger, "ETAPA 3: Archivo ya descomprimido hoy. Se omite descompresion.")

    phase_header(logger, "LECTURA CSV")
    df = leer_csv(logger)
    if 'int_art' not in df.columns:
        raise RuntimeError("No existe columna int_art")
    if 'sales check' not in df.columns:
        raise RuntimeError("No existe columna sales check")

    lista_int_art = df['int_art'].dropna().astype(str).drop_duplicates().tolist()
    total = len(lista_int_art)
    log_console(logger, f"Se detectaron {total} int_art únicos para consultar.")

    phase_header(logger, "CONSULTA WMS INVENTARIO")
    user_wms, pass_wms = cargar_credenciales(CONFIG_WMS)
    auth_wms = (user_wms, pass_wms)

    inventario_phase_done = (
        control.data.get('_inventario_phase_done') is True
        and control.data.get('_inventario_phase_done_fecha') == hoy
        and int(control.data.get('_inventario_total', -1)) == int(total)
    )

    consultados = {k for k in control.consultados() if not str(k).startswith('_')}
    pendientes = [sku for sku in lista_int_art if sku not in consultados]
    log_console(logger, f"Total a consultar: {total}. Ya consultados: {len(consultados)}. Pendientes: {len(pendientes)}.")

    resultados_inv = {k: v for k, v in control.data.items() if not str(k).startswith('_')}
    timeout_list = control.data.get('_timeout_list', [])

    inventario_ejecutado_esta_corrida = False

    if inventario_phase_done:
        log_console(logger, "Fase de inventario ya completada hoy. Se omite consulta API WMS.")
        pendientes = []
        timeout_list = []

    if pendientes:
        inventario_ejecutado_esta_corrida = True
        progreso = {'actual': 0}
        lock = __import__('threading').Lock()
        control_lock = __import__('threading').Lock()

        def worker(sku):
            sku_r, existencia, estatus = consultar_item_wms(sku, auth_wms)
            with control_lock:
                control.registrar(sku_r, existencia)
            resultados_inv[sku_r] = existencia
            with lock:
                progreso['actual'] += 1
                n = progreso['actual']
                print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}...[{n} de {len(pendientes)}] Hecho", flush=True)
                if estatus == 'TIEMPO AGOTADO' and sku_r not in timeout_list:
                    timeout_list.append(sku_r)
                    control.data['_timeout_list'] = timeout_list
                    control.guardar()

        for i in range(0, len(pendientes), 100):
            segmento = pendientes[i:i+100]
            with ThreadPoolExecutor(max_workers=5) as ex:
                list(ex.map(worker, segmento))
            if i + 100 < len(pendientes):
                ini = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                log_console(logger, f"Sleep 1 minuto INICIO: {ini}")
                time.sleep(60)
                fin = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                log_console(logger, f"Sleep 1 minuto FIN: {fin}")

    if timeout_list:
        inventario_ejecutado_esta_corrida = True
        phase_header(logger, "REINTENTO TIMEOUT INVENTARIO")
        ini = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_console(logger, f"Sleep 2 minutos INICIO: {ini}")
        time.sleep(120)
        fin = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_console(logger, f"Sleep 2 minutos FIN: {fin}")

        retry_progress = {'actual': 0}
        retry_lock = __import__('threading').Lock()
        retry_ok = []

        def retry_worker(sku):
            sku_r, existencia, estatus = consultar_item_wms(sku, auth_wms)
            control.registrar(sku_r, existencia)
            resultados_inv[sku_r] = existencia
            with retry_lock:
                retry_progress['actual'] += 1
                n = retry_progress['actual']
                print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}...[{n} de {len(timeout_list)}] Hecho", flush=True)
                if estatus != 'TIEMPO AGOTADO':
                    retry_ok.append(sku_r)

        for i in range(0, len(timeout_list), 100):
            segmento = timeout_list[i:i+100]
            with ThreadPoolExecutor(max_workers=5) as ex:
                list(ex.map(retry_worker, segmento))
            if i + 100 < len(timeout_list):
                ini2 = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                log_console(logger, f"Sleep 1 minuto INICIO: {ini2}")
                time.sleep(60)
                fin2 = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                log_console(logger, f"Sleep 1 minuto FIN: {fin2}")

        control.data['_timeout_list'] = [x for x in timeout_list if x not in retry_ok]
        control.guardar()

    if not inventario_phase_done:
        control.data['_inventario_phase_done'] = True
        control.data['_inventario_phase_done_fecha'] = hoy
        control.data['_inventario_total'] = int(total)
        control.data['_inventario_phase_done_ts'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        control.guardar()
        if inventario_ejecutado_esta_corrida:
            log_console(logger, "Fase de inventario marcada como COMPLETADA para hoy.")

    phase_header(logger, "INSERCION EXISTENCIA WMS")
    if 'Existencia WMS' in df.columns:
        df.drop(columns=['Existencia WMS'], inplace=True)
    if 'existencia' in df.columns:
        idx = df.columns.get_loc('existencia') + 1
    else:
        idx = len(df.columns)
    df.insert(idx, 'Existencia WMS', df['int_art'].astype(str).map(resultados_inv).fillna(0))

    phase_header(logger, "FLUJO RELICES")
    relices_phase_done = (
        control.data.get('_relices_phase_done') is True
        and control.data.get('_relices_phase_done_fecha') == hoy
    )

    if not relices_phase_done:
        relices_txt_done = (
            control.data.get('_relices_txt_done') is True
            and control.data.get('_relices_txt_done_fecha') == hoy
            and os.path.exists(ARCHIVO_RELICES)
        )
        if not relices_txt_done:
            crear_relices_txt(logger, df)
            control.data['_relices_txt_done'] = True
            control.data['_relices_txt_done_fecha'] = hoy
            control.guardar()
        else:
            log_console(logger, "Checkpoint detectado: relices.txt ya generado.")

        relices_ftp_done = (
            control.data.get('_relices_ftp_done') is True
            and control.data.get('_relices_ftp_done_fecha') == hoy
        )
        if not relices_ftp_done:
            if not exists_hidden_checkpoint('relices_ftp_subido'):
                ftp_subir(logger, ARCHIVO_RELICES, FTP_DIR_RELICES, ARCHIVO_RELICES_REMOTO)
                write_hidden_checkpoint('relices_ftp_subido', 'ok')
            else:
                log_console(logger, "Checkpoint detectado: FTP relices ya subido.")
            control.data['_relices_ftp_done'] = True
            control.data['_relices_ftp_done_fecha'] = hoy
            control.guardar()
        else:
            log_console(logger, "Fase FTP de relíces ya completada hoy.")

        relices_ssh_done = (
            control.data.get('_relices_ssh_done') is True
            and control.data.get('_relices_ssh_done_fecha') == hoy
        )
        if not relices_ssh_done:
            if not exists_hidden_checkpoint('relices_ssh_ejecutado'):
                ejecutar_menu_gnx(logger)
                write_hidden_checkpoint('relices_ssh_ejecutado', 'ok')
            else:
                log_console(logger, "Checkpoint detectado: SSH relizador ya ejecutado.")
            control.data['_relices_ssh_done'] = True
            control.data['_relices_ssh_done_fecha'] = hoy
            control.guardar()
        else:
            log_console(logger, "Fase SSH de relíces ya completada hoy.")

        relices_descarga_done = (
            control.data.get('_relices_descarga_done') is True
            and control.data.get('_relices_descarga_done_fecha') == hoy
            and os.path.exists(ARCHIVO_RELICES_PROCESADO)
        )
        if not relices_descarga_done:
            if not exists_hidden_checkpoint('relices_procesado_descargado'):
                ok_proc = esperar_y_descargar_relices_procesado(logger, max_espera_seg=1800, cada_seg=30)
                if not ok_proc:
                    raise RuntimeError("No se pudo obtener relices_procesado.txt en tiempo esperado")
                write_hidden_checkpoint('relices_procesado_descargado', 'ok')
            else:
                log_console(logger, "Checkpoint detectado: relices_procesado ya descargado.")
                if not os.path.exists(ARCHIVO_RELICES_PROCESADO):
                    ok_proc2 = esperar_y_descargar_relices_procesado(logger, max_espera_seg=600, cada_seg=30)
                    if not ok_proc2:
                        raise RuntimeError("Checkpoint existe pero no está relices_procesado.txt local")
            control.data['_relices_descarga_done'] = True
            control.data['_relices_descarga_done_fecha'] = hoy
            control.guardar()
        else:
            log_console(logger, "Fase descarga de relices_procesado ya completada hoy.")

        mapeo_sales_order = parsear_relices_procesado(logger)
        if not mapeo_sales_order:
            control.data['_relices_phase_done'] = False
            control.data['_relices_phase_done_fecha'] = hoy
            control.guardar()
            raise RuntimeError("relices_procesado.txt no generó mapeo válido (0 registros). No se puede continuar a order_hdr.")
        control.data['_mapeo_sales_order'] = mapeo_sales_order
        control.data['_relices_phase_done'] = True
        control.data['_relices_phase_done_fecha'] = hoy
        control.guardar()
        log_console(logger, "Fase relíces marcada como COMPLETADA para hoy.")
    else:
        log_console(logger, "Fase relíces ya completada hoy. Se reutiliza información guardada.")
        mapeo_sales_order = control.data.get('_mapeo_sales_order', {})
        if not mapeo_sales_order:
            mapeo_sales_order = parsear_relices_procesado(logger)
            if not mapeo_sales_order:
                raise RuntimeError("No hay mapeo de relices disponible (0 registros). Revisa relices_procesado.txt")
            control.data['_mapeo_sales_order'] = mapeo_sales_order
            control.guardar()

    mapeo_sales_order_nocero = {strip_leading_zero(k): v for k, v in mapeo_sales_order.items()}

    order_unicos = list(dict.fromkeys([str(v).strip() for v in mapeo_sales_order.values() if str(v).strip()]))
    if not order_unicos:
        raise RuntimeError("No se encontraron order_nbr válidos para consultar en order_hdr (total objetivo = 0).")
    order_status_map = control.data.get('_order_status_map', {})
    order_timeout_list = control.data.get('_order_timeout_list', [])
    order_no_encontrada_list = control.data.get('_order_no_encontrada_list', [])

    order_unicos_objetivo = list(order_unicos)

    order_phase_done = (
        control.data.get('_order_phase_done') is True
        and control.data.get('_order_phase_done_fecha') == hoy
        and int(control.data.get('_order_total', -1)) == int(len(order_unicos))
    )

    if not order_phase_done:
        order_timeout_list = [x for x in order_timeout_list if x in order_unicos_objetivo]
        order_no_encontrada_list = [x for x in order_no_encontrada_list if x in order_unicos_objetivo]
        order_pendientes = [x for x in order_unicos_objetivo if x not in order_status_map]
        order_ya_consultadas = len(order_unicos_objetivo) - len(order_pendientes)
        log_console(
            logger,
            f"order_hdr total objetivo: {len(order_unicos_objetivo)}. Ya consultadas: {order_ya_consultadas}. Pendientes: {len(order_pendientes)}."
        )

        total_lotes_order = (len(order_pendientes) + 99) // 100 if order_pendientes else 0
        for i in range(0, len(order_pendientes), 100):
            lote = order_pendientes[i:i+100]
            lote_n = (i // 100) + 1
            log_console(logger, f"order_hdr lote {lote_n}/{total_lotes_order} ({len(lote)} órdenes)")
            lote_status_map, _ = consultar_order_status_lote(lote, auth_wms)

            for order_nbr in lote:
                estatus = lote_status_map.get(order_nbr, 'No Encontrada')
                order_status_map[order_nbr] = estatus
                if estatus == 'TIEMPO AGOTADO' and order_nbr not in order_timeout_list:
                    order_timeout_list.append(order_nbr)
                if estatus != 'TIEMPO AGOTADO' and order_nbr in order_timeout_list:
                    order_timeout_list.remove(order_nbr)
                if estatus == 'No Encontrada' and order_nbr not in order_no_encontrada_list:
                    order_no_encontrada_list.append(order_nbr)
                if estatus != 'No Encontrada' and order_nbr in order_no_encontrada_list:
                    order_no_encontrada_list.remove(order_nbr)

            control.data['_order_status_map'] = order_status_map
            control.data['_order_timeout_list'] = order_timeout_list
            control.data['_order_no_encontrada_list'] = order_no_encontrada_list
            control.guardar()

        def preparar_reintento_objetivo():
            rt = list(dict.fromkeys([x for x in order_timeout_list if x in order_unicos_objetivo]))
            rn = list(dict.fromkeys([x for x in order_no_encontrada_list if x in order_unicos_objetivo]))
            ru = list(dict.fromkeys(rt + rn))
            return rt, rn, ru

        def ejecutar_reintento(stage_name, sleep_min, objetivo, modo='lote'):
            nonlocal order_timeout_list, order_no_encontrada_list
            if not objetivo:
                return

            log_console(logger, "Cerrando conexiones HTTP de order_hdr.")
            ini = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_console(logger, f"Sleep {sleep_min} minutos INICIO: {ini}")
            time.sleep(sleep_min * 60)
            fin = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            log_console(logger, f"Sleep {sleep_min} minutos FIN: {fin}")

            timeout_restante = []
            no_encontrada_restante = []
            cambiaron_estado = 0
            permanecieron_igual = 0

            if modo in ('lote', 'lote30', 'lote20', 'lote10'):
                if modo == 'lote':
                    tam_lote = 100
                elif modo == 'lote30':
                    tam_lote = 30
                elif modo == 'lote20':
                    tam_lote = 20
                else:
                    tam_lote = 10
                total_lotes_retry = (len(objetivo) + tam_lote - 1) // tam_lote if objetivo else 0
                for i in range(0, len(objetivo), tam_lote):
                    lote = objetivo[i:i+tam_lote]
                    lote_n = (i // tam_lote) + 1
                    log_console(logger, f"{stage_name} lote {lote_n}/{total_lotes_retry} ({len(lote)} órdenes)")
                    lote_status_map, _ = consultar_order_status_lote(lote, auth_wms)

                    for order_nbr in lote:
                        estatus_previo = order_status_map.get(order_nbr, '')
                        estatus = lote_status_map.get(order_nbr, 'No Encontrada')
                        order_status_map[order_nbr] = estatus
                        if estatus_previo == estatus:
                            permanecieron_igual += 1
                        else:
                            cambiaron_estado += 1
                        if estatus == 'TIEMPO AGOTADO':
                            timeout_restante.append(order_nbr)
                        elif estatus == 'No Encontrada':
                            no_encontrada_restante.append(order_nbr)

                    control.data['_order_status_map'] = order_status_map
                    control.data['_order_timeout_list'] = timeout_restante
                    control.data['_order_no_encontrada_list'] = no_encontrada_restante
                    control.guardar()
            else:
                for i, order_nbr in enumerate(objetivo, 1):
                    log_console(logger, f"{stage_name} {i}/{len(objetivo)} (1 orden)")
                    estatus_previo = order_status_map.get(order_nbr, '')
                    estatus = consultar_order_status_unitario(order_nbr, auth_wms)
                    order_status_map[order_nbr] = estatus
                    if estatus_previo == estatus:
                        permanecieron_igual += 1
                    else:
                        cambiaron_estado += 1
                    if estatus == 'TIEMPO AGOTADO':
                        timeout_restante.append(order_nbr)
                    elif estatus == 'No Encontrada':
                        no_encontrada_restante.append(order_nbr)

                    control.data['_order_status_map'] = order_status_map
                    control.data['_order_timeout_list'] = timeout_restante
                    control.data['_order_no_encontrada_list'] = no_encontrada_restante
                    control.guardar()

            order_timeout_list = timeout_restante
            order_no_encontrada_list = no_encontrada_restante
            control.data['_order_timeout_list'] = order_timeout_list
            control.data['_order_no_encontrada_list'] = order_no_encontrada_list
            control.guardar()

            log_console(
                logger,
                f"Resultado {stage_name} -> Total revisadas: {len(objetivo)} | Cambiaron estado: {cambiaron_estado} | Permanecieron igual: {permanecieron_igual}"
            )

        rt1, rn1, ru1 = preparar_reintento_objetivo()
        if ru1:
            phase_header(logger, "REINTENTO 1 TIMEOUT/NO ENCONTRADA ORDER_HDR")
            log_console(
                logger,
                f"Reconsulta 1 tras sleep -> TIEMPO AGOTADO: {len(rt1)} | No Encontrada: {len(rn1)} | Total único: {len(ru1)}"
            )
            ejecutar_reintento("order_hdr retry1", 2, ru1, modo='lote30')

        rt2, rn2, ru2 = preparar_reintento_objetivo()
        if ru2:
            phase_header(logger, "REINTENTO 2 TIMEOUT/NO ENCONTRADA ORDER_HDR")
            log_console(
                logger,
                f"Reconsulta 2 tras sleep -> TIEMPO AGOTADO: {len(rt2)} | No Encontrada: {len(rn2)} | Total único: {len(ru2)}"
            )
            ejecutar_reintento("order_hdr retry2", 2, ru2, modo='lote20')

        rt3, rn3, ru3 = preparar_reintento_objetivo()
        if ru3:
            phase_header(logger, "REINTENTO 3 FINAL LOTES 10 ORDER_HDR")
            log_console(
                logger,
                f"Reconsulta 3 tras sleep -> TIEMPO AGOTADO: {len(rt3)} | No Encontrada: {len(rn3)} | Total único: {len(ru3)}"
            )
            ejecutar_reintento("order_hdr retry3", 2, ru3, modo='lote10')

        control.data['_order_status_map'] = order_status_map
        all_objetivo_resuelto = all(
            (x in order_status_map) and (order_status_map.get(x) not in ('TIEMPO AGOTADO', 'No Encontrada'))
            for x in order_unicos_objetivo
        )

        if all_objetivo_resuelto and not order_timeout_list and not order_no_encontrada_list:
            control.data['_order_phase_done'] = True
            control.data['_order_phase_done_fecha'] = hoy
            control.data['_order_total'] = int(len(order_unicos))
        else:
            control.data['_order_phase_done'] = False
            control.data['_order_phase_done_fecha'] = hoy
            control.data['_order_total'] = int(len(order_unicos))
        control.guardar()
        if all_objetivo_resuelto and not order_timeout_list and not order_no_encontrada_list:
            log_console(logger, "Fase order_hdr marcada como COMPLETADA para hoy.")
        elif order_timeout_list or order_no_encontrada_list:
            log_console(
                logger,
                f"Fase order_hdr parcial: quedaron TIEMPO AGOTADO={len(order_timeout_list)} | No Encontrada={len(order_no_encontrada_list)} para próximas corridas."
            )
    else:
        log_console(logger, "Fase order_hdr ya completada hoy. Se omiten consultas.")

    phase_header(logger, "INSERCION ULTIMA ORDEN / ESTATUS WMS")
    if 'Ultima orden' in df.columns:
        df.drop(columns=['Ultima orden'], inplace=True)
    if 'Estatus WMS' in df.columns:
        df.drop(columns=['Estatus WMS'], inplace=True)

    idx_wms = df.columns.get_loc('Existencia WMS') + 1
    df.insert(idx_wms, 'Ultima orden', '')
    df.insert(idx_wms + 1, 'Estatus WMS', '')

    ultimas = []
    estatuses = []
    for sales in df['sales check'].tolist():
        sales_norm = normalize_sales_check(sales)
        ord_rel = mapeo_sales_order.get(sales_norm)
        if not ord_rel:
            ord_rel = mapeo_sales_order_nocero.get(strip_leading_zero(sales_norm))
        if ord_rel:
            ultimas.append(ord_rel)
            estatuses.append(order_status_map.get(ord_rel, 'No Encontrada'))
        else:
            ultimas.append('')
            estatuses.append('')

    df['Ultima orden'] = ultimas
    df['Estatus WMS'] = estatuses

    phase_header(logger, "GENERACION EXCEL")
    archivo_final = guardar_excel(df, logger)
    log_console(logger, "Archivo Excel Generado Exitosamente")
    log_only(logger, f"Archivo Excel generado: {archivo_final}")
    control.data['_excel_generado'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    control.data['_excel_phase_done'] = True
    control.data['_excel_phase_done_fecha'] = hoy
    control.guardar()

    ejecutar_fase_envio_correo(logger, control, hoy, archivo_final)

    log_only(logger, f"=== PROCESO FINALIZADO CORRECTAMENTE: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")


# ®Carlos Alfonso Ortega Molina®
if __name__ == "__main__":
    main()
